"""Записи, клиенты и Excel-расписание для BARBERVAULT."""
import csv
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BOOKINGS_PATH = os.path.join(BASE_DIR, "bookings.csv")
CLIENTS_PATH = os.path.join(BASE_DIR, "clients.csv")
SCHEDULE_PATH = os.path.join(BASE_DIR, "weekly_schedule.xlsx")

BOOKING_COLUMNS = [
    "id", "date", "time", "service", "price", "duration_min", "barber",
    "telegram_id", "client_name", "phone", "email", "booked_at",
]
CLIENT_COLUMNS = [
    "telegram_id", "client_name", "phone", "email", "privacy_accepted_at",
    "privacy_version", "created_at", "updated_at",
]
TIME_SLOTS = [f"{hour:02d}:00" for hour in range(10, 20)]
WORKDAY_END_MINUTES = 20 * 60


def _ensure_csv(path: str, columns: list[str]) -> None:
    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            csv.writer(fh).writerow(columns)


def _read_csv(path: str, columns: list[str]) -> list[dict[str, str]]:
    _ensure_csv(path, columns)
    with open(path, encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: str, columns: list[str], rows: list[dict[str, str]]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def load_bookings() -> list[dict[str, str]]:
    return _read_csv(BOOKINGS_PATH, BOOKING_COLUMNS)


def get_client(telegram_id: int) -> dict[str, str] | None:
    telegram_id_str = str(telegram_id)
    return next(
        (row for row in _read_csv(CLIENTS_PATH, CLIENT_COLUMNS)
         if row.get("telegram_id") == telegram_id_str),
        None,
    )


def save_client(client: dict) -> None:
    rows = _read_csv(CLIENTS_PATH, CLIENT_COLUMNS)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    telegram_id = str(client["telegram_id"])
    new_row = {
        "telegram_id": telegram_id,
        "client_name": client["client_name"],
        "phone": client["phone"],
        "email": client.get("email", ""),
        "privacy_accepted_at": client.get("privacy_accepted_at", now),
        "privacy_version": client.get("privacy_version", "2026-08-03"),
        "created_at": now,
        "updated_at": now,
    }
    for index, row in enumerate(rows):
        if row.get("telegram_id") == telegram_id:
            new_row["created_at"] = row.get("created_at") or now
            rows[index] = new_row
            break
    else:
        rows.append(new_row)
    _write_csv(CLIENTS_PATH, CLIENT_COLUMNS, rows)


def _minutes(time_str: str) -> int:
    hour, minute = map(int, time_str.split(":"))
    return hour * 60 + minute


def _booking_duration(booking: dict[str, str]) -> int:
    try:
        return int(booking.get("duration_min", "60"))
    except ValueError:
        return 60


def is_slot_available(date_str: str, time_str: str, barber: str, duration_min: int) -> bool:
    """Проверяет пересечение интервалов только у выбранного мастера."""
    start = _minutes(time_str)
    end = start + duration_min
    if end > WORKDAY_END_MINUTES:
        return False

    for booking in load_bookings():
        if booking.get("date") != date_str or booking.get("barber") != barber:
            continue
        booking_start = _minutes(booking["time"])
        booking_end = booking_start + _booking_duration(booking)
        if start < booking_end and booking_start < end:
            return False
    return True


def unavailable_slots(date_str: str, barber: str, duration_min: int) -> set[str]:
    return {
        time_str for time_str in TIME_SLOTS
        if not is_slot_available(date_str, time_str, barber, duration_min)
    }


def save_booking(booking: dict) -> None:
    rows = load_bookings()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows.append({
        "id": str(len(rows) + 1),
        "date": booking["date"],
        "time": booking["time"],
        "service": booking["service"],
        "price": str(booking["price"]),
        "duration_min": str(booking["duration_min"]),
        "barber": booking["barber"],
        "telegram_id": str(booking["telegram_id"]),
        "client_name": booking["client_name"],
        "phone": booking["phone"],
        "email": booking.get("email", ""),
        "booked_at": now,
    })
    _write_csv(BOOKINGS_PATH, BOOKING_COLUMNS, rows)
    generate_weekly_schedule()


def generate_weekly_schedule(start_date: datetime | None = None) -> None:
    """Создаёт Excel-файл на семь ближайших календарных дней, по листу на мастера."""
    from data import BARBERS, WEEKDAYS

    start = (start_date or datetime.now()).date()
    dates = [start + timedelta(days=offset) for offset in range(7)]
    bookings = load_bookings()

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    free_fill = PatternFill("solid", fgColor="E2F0D9")
    busy_fill = PatternFill("solid", fgColor="FCE4D6")

    for barber in BARBERS.values():
        sheet = workbook.create_sheet(title=barber["name"])
        sheet.append(["Время", *[f"{day:%d.%m} ({WEEKDAYS[day.weekday()]})" for day in dates]])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for time_str in TIME_SLOTS:
            row = [time_str]
            for day in dates:
                date_str = day.isoformat()
                matching = []
                slot_start = _minutes(time_str)
                slot_end = slot_start + 60
                for booking in bookings:
                    if booking.get("date") != date_str or booking.get("barber") != barber["name"]:
                        continue
                    start_minutes = _minutes(booking["time"])
                    end_minutes = start_minutes + _booking_duration(booking)
                    if slot_start < end_minutes and start_minutes < slot_end:
                        matching.append(booking)
                if matching:
                    booking = matching[0]
                    row.append(f"ЗАНЯТО\n{booking['service']}\n{booking['client_name']}")
                else:
                    row.append("СВОБОДНО")
            sheet.append(row)

        sheet.column_dimensions["A"].width = 12
        for column in range(2, 9):
            sheet.column_dimensions[chr(64 + column)].width = 26
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = busy_fill if str(cell.value).startswith("ЗАНЯТО") else free_fill
        for row_number in range(2, 12):
            sheet.row_dimensions[row_number].height = 48
        sheet.freeze_panes = "B2"

    workbook.save(SCHEDULE_PATH)
